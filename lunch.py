#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import easygui
import warnings
import os
import datetime


#from __future__ import print_function

#Ignore warnings
warnings.filterwarnings("ignore")
#----------------------VARIABLES----------------------------
#MODDED FONTE IN A list of lists
salasFonte = [[5,6,8,9,10],
	[12,13,14,15,16,17],
	[19,20,21,22,23,24],
	[26,27,28,29,30,31],
	[33,34,35,36]]
letrasF = 	['B','E','H','K','N']
#salas = 	[s1BerF,s2BerF,sParqueF,sPreCatlF,dietaF]
	

#TARGET
ementaNrTgt = "C2"
salasTarget = [[6,7,9,10,11], 
	[15,16,17,18,19,20],	
	[24,25,26,27,28,29],	
	[33,34,35,36,37,38],	
	[42,43,44,45]]			
letrasT = 	['B','C','D','E','F']

posSegunda = "J3"
posSexta = "J4"
#salas = 	[s1BerT,s2BerT,sParqueT,sPreCatlT,dietaT]

#format dia
fExl ="%d-%m-%Y"
fFile ="%d_%m_%Y"
#Nr Ementa
ementaNr = easygui.integerbox(msg="Qual é o número da ementa?",
title ="Ementa",default=0,lowerbound=0,upperbound=10)

#Dia de hoje
hoje = datetime.date.today()
hojeF = hoje.strftime("%d_%m_%Y")
proxSeg = 0
proxSexta = 0
#----------------END OF VARIABLES------------------


#Ficheiro fonte
ficheiroFonte = easygui.fileopenbox()

#Ficheiro a gravar
ficheiroTarget = os.path.join(os.getcwd(), "modelo.xlsx")

if ficheiroFonte != None and ficheiroTarget != False :
	#Workbook source
	wbFonte = load_workbook(ficheiroFonte)
	wsFonte = wbFonte.worksheets[0]

	#Workbook target
	wbTarget = load_workbook(ficheiroTarget)
	wsTarget = wbTarget.worksheets[0]


#Este método calcula a próxima segunda e sexta feira
def calcularDia():
	global proxSeg
	global proxSexta
	if datetime.date.today() != None:
		x = datetime.date.today()
		while 1:
			if datetime.date.today().weekday() == 0:
				x =  datetime.date.today() + datetime.timedelta(1)
			else:
				x += datetime.timedelta(1)
			if x.weekday() == 0:
				proxSeg = x
				proxSexta = x + datetime.timedelta(4)
				break
	return

#Este método copia dados um ficheiro de excel para outro
def copiarDadosExl():
	for lF in range(len(letrasF)):
		for nrS in range(len(salasFonte)):
				for nrL in range(len(salasFonte[nrS])):
					wsTarget["%s%d" % (letrasT[lF],salasTarget[nrS][nrL])].value = wsFonte["%s%d" % (letrasF[lF],salasFonte[nrS][nrL])].value


calcularDia()
copiarDadosExl()

wsTarget[posSegunda].value = proxSeg.strftime(fExl)
wsTarget[posSexta].value = proxSexta.strftime(fExl)
wsTarget[ementaNrTgt].value = ementaNr
wbTarget.save("%s_%s_%s%s" %("resultado",ementaNr,hojeF,".xlsx"))
